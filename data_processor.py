import openpyxl
import pandas as pd
import sys
import os
import datetime
import re
from copy import copy

def clean_numeric(val):
    """
    Robustly cleans numeric values from Excel strings/numbers.
    Handles Vietnamese diacritics, spaces, thousands separators (dots/spaces),
    and commas for decimals.
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    
    s = str(val).replace('\xa0', '').replace(' ', '').strip()
    if not s:
        return 0.0
    
    # If there is a comma, it represents a decimal point in Vietnamese format (e.g. 10 286 294,00)
    # We remove dots (which are thousands separators) and replace commas with dots (decimals)
    if ',' in s:
        s = s.replace('.', '')
        s = s.replace(',', '.')
    
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_import(file_path):
    """
    Parses the import spreadsheet (INV-007A).
    Returns a list of parsed import transaction records.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    current_voucher = None
    current_date = None
    current_desc = None
    
    records = []
    
    for r in range(11, sheet.max_row + 1):
        stt = sheet.cell(row=r, column=1).value
        v_code = sheet.cell(row=r, column=2).value
        v_date = sheet.cell(row=r, column=3).value
        v_desc = sheet.cell(row=r, column=7).value
        m_code = sheet.cell(row=r, column=8).value
        
        # Detect Header Row (STT is not None, Voucher Code is not None, Material Code is None)
        if stt is not None and v_code is not None and m_code is None:
            current_voucher = str(v_code).strip()
            current_date = v_date
            current_desc = str(v_desc).strip() if v_desc is not None else ""
            continue
            
        # Detect Detail Row (Material Code is not None)
        if m_code is not None:
            m_code_str = str(m_code).strip()
            if m_code_str.lower() in ["cộng", "tổng cộng", "tổng"]:
                continue
                
            m_name = sheet.cell(row=r, column=9).value
            m_unit = sheet.cell(row=r, column=13).value
            m_qty = sheet.cell(row=r, column=14).value
            m_price = sheet.cell(row=r, column=15).value
            m_amount = sheet.cell(row=r, column=17).value
            
            qty_val = clean_numeric(m_qty)
            price_val = clean_numeric(m_price)
            amount_val = clean_numeric(m_amount)
            
            records.append({
                "type": "NHAP",
                "voucher": current_voucher,
                "date": current_date,
                "desc": current_desc,
                "code": m_code_str,
                "name": str(m_name).strip() if m_name else "",
                "unit": str(m_unit).strip() if m_unit else "",
                "qty": qty_val,
                "price": price_val,
                "amount": amount_val
            })
            
    return records

def parse_export(file_path):
    """
    Parses the export spreadsheet (INV-009).
    Returns a list of parsed export transaction records.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    current_voucher_02 = None
    current_voucher_03 = None
    current_date = None
    current_desc = None
    
    records = []
    
    for r in range(11, sheet.max_row + 1):
        stt = sheet.cell(row=r, column=1).value
        v_code = sheet.cell(row=r, column=2).value
        v_date = sheet.cell(row=r, column=3).value
        v_desc = sheet.cell(row=r, column=5).value
        m_code = sheet.cell(row=r, column=5).value # Material Code is shifted to Column E (5)
        m_name = sheet.cell(row=r, column=7).value # Material Name is shifted to Column G (7)
        
        # Detect Header Row 1 (contains STT and 02.VH... code)
        if stt is not None and v_code is not None and str(v_code).strip().startswith("02"):
            current_voucher_02 = str(v_code).strip()
            current_desc = str(v_desc).strip() if v_desc is not None else ""
            current_voucher_03 = None
            current_date = None
            continue
            
        # Detect Header Row 2 (contains 03.VH... code and transaction date)
        if stt is None and v_code is not None and str(v_code).strip().startswith("03"):
            current_voucher_03 = str(v_code).strip()
            current_date = v_date
            continue
            
        # Detect Detail Row (STT is None, Voucher No is None, Material Code is not None)
        if stt is None and v_code is None and m_code is not None:
            m_code_str = str(m_code).strip()
            if m_code_str.lower().startswith("mục đích") or m_code_str.lower() in ["cộng", "tổng cộng", "tổng"]:
                continue
                
            m_unit = sheet.cell(row=r, column=11).value # Column K is Unit
            m_qty = sheet.cell(row=r, column=13).value # Column M is Quantity
            m_price = sheet.cell(row=r, column=14).value # Column N is Unit Price
            m_amount = sheet.cell(row=r, column=15).value # Column O is Amount
            
            qty_val = clean_numeric(m_qty)
            price_val = clean_numeric(m_price)
            amount_val = clean_numeric(m_amount)
            
            # For exports, openpyxl reads raw quantities multiplied by 1000.
            # We divide raw numeric quantities by 1000 to get the correct standard amount.
            if isinstance(m_qty, (int, float)):
                qty_val = qty_val / 1000.0
                
            # Combine voucher codes: voucher_02 + voucher_03
            combined_voucher = ""
            if current_voucher_02 and current_voucher_03:
                combined_voucher = f"{current_voucher_02}{current_voucher_03}"
            elif current_voucher_02:
                combined_voucher = current_voucher_02
            elif current_voucher_03:
                combined_voucher = current_voucher_03
                
            records.append({
                "type": "XUAT",
                "voucher": combined_voucher,
                "date": current_date,
                "desc": current_desc,
                "code": m_code_str,
                "name": str(m_name).strip() if m_name else "",
                "unit": str(m_unit).strip() if m_unit else "",
                "qty": qty_val,
                "price": price_val,
                "amount": amount_val
            })
            
    return records

def classify_import_type(desc, code=None):
    """
    Classifies an Import (NHAP) record:
    - 'HOAN_TRA': Nhập lại / hoàn nhập vật tư thừa chưa sử dụng hoàn trả kho
    - 'VTTH': Nhập vật tư thu hồi (tháo gỡ, thay thế, xác/phế liệu)
    """
    desc_str = str(desc or "").upper()
    code_str = str(code or "").upper()
    
    hoan_tra_keywords = ["HOÀN NHẬP", "HOAN NHAP", "NHẬP LẠI", "NHAP LAI", "TRẢ LẠI", "TRA LAI", "KHÔNG SỬ DỤNG", "KHONG SU DUNG", "HOÀN TRẢ", "HOAN TRA"]
    if any(k in desc_str for k in hoan_tra_keywords):
        return "HOAN_TRA"
        
    vtth_keywords = ["THU HỒI", "THU HOI", "VTTH", "THÁO GỠ", "THAO GO", "THANH LÝ", "THANH LY", "XÁC", "PHẾ LIỆU"]
    if any(k in desc_str for k in vtth_keywords):
        return "VTTH"
        
    if code_str.endswith((".CXA", ".BXX")):
        return "VTTH"
        
    return "VTTH" if "THU HỒI" in desc_str else "HOAN_TRA"

def consolidate_data(import_records, export_records, scl_only=True, warehouse_filter=None, keyword_filter=None):
    """
    Consolidates parsed import and export records into a single formatted list.
    Supports filtering by SCL keyword, warehouse prefix, and general keyword search.
    Returns a pandas DataFrame sorted by transaction date.
    """
    rows = []
    
    # Process Imports
    for r in import_records:
        wh = ""
        if r["voucher"] and len(r["voucher"].split('.')) > 1:
            wh = r["voucher"].split('.')[1]
            
        # Filters (Space-insensitive check for SCL and case-insensitive check for VTAD)
        if scl_only:
            desc_clean = r["desc"].replace(" ", "").upper() if r["desc"] else ""
            voucher_clean = r["voucher"].replace(" ", "").upper() if r["voucher"] else ""
            if "SCL" not in desc_clean and "SCL" not in voucher_clean and "VTAD" not in desc_clean and "VTAD" not in voucher_clean and "VTDA" not in desc_clean and "VTDA" not in voucher_clean:
                continue
        if warehouse_filter and wh not in warehouse_filter:
            continue
        if keyword_filter:
            kw = keyword_filter.upper()
            if kw not in r["code"].upper() and kw not in r["name"].upper() and kw not in r["desc"].upper():
                continue
                
        # Parse Month and Year
        m = None
        y = None
        if isinstance(r["date"], (datetime.datetime, datetime.date)):
            m = r["date"].month
            y = r["date"].year
            
        itype = classify_import_type(r["desc"], r["code"])
        type_label = "Nhập - Hoàn Trả (Vật tư thừa)" if itype == "HOAN_TRA" else "Nhập - VTTH (Thu hồi tháo gỡ)"
            
        rows.append({
            "Mã vật tư": r["code"],
            "Tên vật tư": r["name"],
            "ĐVT": r["unit"],
            "tháng": m,
            "năm": y,
            "Ngày viết": r["date"],
            "Số chứng từ (Nhập)": r["voucher"],
            "Số chứng từ (Xuất)": None,
            "Diễn giải": r["desc"],
            "Nhập - Số lượng": r["qty"],
            "Nhập - Đơn giá": r["price"],
            "Nhập - Thành tiền": r["amount"],
            "Xuất - Số lượng": 0,
            "Xuất - Đơn giá": 0,
            "Xuất - Thành tiền": 0,
            "Loại giao dịch": type_label
        })
        
    # Process Exports
    for r in export_records:
        wh = ""
        if r["voucher"] and len(r["voucher"].split('.')) > 1:
            wh = r["voucher"].split('.')[1]
            
        # Filters (Space-insensitive check for SCL and case-insensitive check for VTAD)
        if scl_only:
            desc_clean = r["desc"].replace(" ", "").upper() if r["desc"] else ""
            voucher_clean = r["voucher"].replace(" ", "").upper() if r["voucher"] else ""
            if "SCL" not in desc_clean and "SCL" not in voucher_clean and "VTAD" not in desc_clean and "VTAD" not in voucher_clean and "VTDA" not in desc_clean and "VTDA" not in voucher_clean:
                continue
        if warehouse_filter and wh not in warehouse_filter:
            continue
        if keyword_filter:
            kw = keyword_filter.upper()
            if kw not in r["code"].upper() and kw not in r["name"].upper() and kw not in r["desc"].upper():
                continue
                
        # Parse Month and Year
        m = None
        y = None
        if isinstance(r["date"], (datetime.datetime, datetime.date)):
            m = r["date"].month
            y = r["date"].year
            
        rows.append({
            "Mã vật tư": r["code"],
            "Tên vật tư": r["name"],
            "ĐVT": r["unit"],
            "tháng": m,
            "năm": y,
            "Ngày viết": r["date"],
            "Số chứng từ (Nhập)": None,
            "Số chứng từ (Xuất)": r["voucher"],
            "Diễn giải": r["desc"],
            "Nhập - Số lượng": 0,
            "Nhập - Đơn giá": 0,
            "Nhập - Thành tiền": 0,
            "Xuất - Số lượng": r["qty"],
            "Xuất - Đơn giá": r["price"],
            "Xuất - Thành tiền": r["amount"],
            "Loại giao dịch": "Xuất - Phục vụ SCL"
        })
        
    df = pd.DataFrame(rows)
    if not df.empty:
        # Sort chronologically by date
        df["sort_date"] = pd.to_datetime(df["Ngày viết"], errors="coerce")
        df = df.sort_values(by="sort_date", ascending=True).drop(columns=["sort_date"])
        
    return df

def write_to_template(df, template_path, output_path):
    """
    Loads the template sheet, clears old data starting from row 3, 
    and writes the new records while copying styles, fonts, borders, 
    alignments, and number formats from the original row 3.
    """
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 1. Capture the formatting styles from Row 3 to use as a blueprint
    style_blueprint = {}
    for col_idx in range(1, 16):
        cell = sheet.cell(row=3, column=col_idx)
        style_blueprint[col_idx] = {
            "font": copy(cell.font) if cell.font else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "border": copy(cell.border) if cell.border else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format
        }
        
    # 2. Clear old rows (starting at row 3)
    original_max_row = sheet.max_row
    if original_max_row >= 3:
        sheet.delete_rows(3, original_max_row - 2)
        
    # 3. Write new records
    if not df.empty:
        for r_idx, row_tuple in enumerate(df.itertuples(index=False), start=3):
            for c_idx, cell_value in enumerate(row_tuple, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                
                # Format dates nicely
                if isinstance(cell_value, (datetime.datetime, datetime.date)):
                    if isinstance(cell_value, datetime.datetime):
                        cell.value = cell_value.date()
                    else:
                        cell.value = cell_value
                else:
                    cell.value = cell_value
                
                # Apply preserved style blueprint
                style = style_blueprint.get(c_idx)
                if style:
                    if style["font"]: cell.font = style["font"]
                    if style["fill"]: cell.fill = style["fill"]
                    if style["border"]: cell.border = style["border"]
                    if style["alignment"]: cell.alignment = style["alignment"]
                    if style["number_format"]: cell.number_format = style["number_format"]
                    
    try:
        wb.save(output_path)
    except PermissionError:
        print(f"Warning: File '{output_path}' is currently open in another program and cannot be saved.", file=sys.stderr)
    except Exception as e:
        print(f"Error saving '{output_path}': {e}", file=sys.stderr)
    return len(df)


# ==============================================================================
# UPGRADE: VOLTAGE LEVEL SEPARATION & MONTHLY DISTRIBUTION
# ==============================================================================

# Standard standardized project names mapping
PROJECT_NAMES = {
    "VTAD2606001": "VTAD2606001 - Sửa chữa lớn TSCĐ hệ thống đo đếm trên địa bàn Công ty Điện Lực Vũng Tàu năm 2026 - Phần bảo trì TU, TI",
    "VTAD2606002": "VTAD2606002 - Sửa chữa lớn FCO, LA năm 2026",
    "VTAD2605001": "VTAD2605001 - Sửa chữa lớn đường dây trung hạ thế, trạm biến áp năm 2026",
    "VTAD2608001": "VTAD2608001 - Sửa chữa lớn Công xa năm 2026"
}

VOLTAGE_MAPPING_CACHE = None

def load_voltage_mapping(file_path="TachPP_BL mẫu.xlsx"):
    """
    Loads and caches the manual voltage level classifications from TachPP_BL mẫu.xlsx.
    """
    global VOLTAGE_MAPPING_CACHE
    if VOLTAGE_MAPPING_CACHE is not None:
        return VOLTAGE_MAPPING_CACHE
        
    mapping = {}
    resolved_path = file_path
    if not os.path.exists(resolved_path):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        resolved_path = os.path.join(script_dir, "TachPP_BL mẫu.xlsx")
        
    if os.path.exists(resolved_path):
        try:
            wb = openpyxl.load_workbook(resolved_path, data_only=True)
            ws = wb.active
            for r in range(3, ws.max_row + 1):
                code_val = ws.cell(row=r, column=1).value
                class_val = ws.cell(row=r, column=16).value
                if code_val and class_val:
                    code_str = str(code_val).strip()
                    class_str = str(class_val).strip()
                    if class_str.lower() in ["trung thế", "trung thê", "trung the"]:
                        mapping[code_str] = "Trung thế"
                    elif class_str.lower() in ["hạ thế", "hạ thê", "ha the"]:
                        mapping[code_str] = "Hạ thế"
        except Exception as e:
            print(f"Error loading voltage mapping template: {e}", file=sys.stderr)
            
    VOLTAGE_MAPPING_CACHE = mapping
    return VOLTAGE_MAPPING_CACHE

def classify_voltage(code, name, desc):
    """
    Classifies SCL materials into Medium Voltage (Trung thế) or Low Voltage (Hạ thế)
    based on the reference template TachPP_BL mẫu.xlsx if available, or standard rules.
    """
    code_str = str(code).strip()
    
    # 1. Try template mapping first
    mapping = load_voltage_mapping()
    if code_str in mapping:
        return mapping[code_str]
        
    # 2. Heuristics fallback
    name_upper = str(name).upper()
    desc_upper = str(desc).upper()
    
    # Medium Voltage (Trung thế)
    # Starts with 3.53.60 (Medium current transformer Gelex), 3.53.65 (Medium current transformer Mitex), 
    # or 3.56 (Voltage transformers TU)
    if code_str.startswith(("3.53.60", "3.53.65", "3.56")):
        return "Trung thế"
        
    # Voltage transformer voltage ratings (22kV, 12kV...)
    mv_indicators = ["24KV", "22KV", "12KV", "110KV", "35KV", "TRUNG THẾ", "TRUNG THÊ", "22:√3", "12000/120V", "22(15):V3"]
    if any(k in name_upper or k in desc_upper for k in mv_indicators):
        return "Trung thế"
        
    # Low Voltage (Hạ thế)
    # Starts with 3.53.05 (Biến dòng hạ thế), 3.53.08 (Máy biến dòng Gelex 600V)
    if code_str.startswith(("3.53.05", "3.53.08")):
        return "Hạ thế"
        
    lv_indicators = ["600V", "720V", "0.4KV", "HẠ THẾ", "HẠ THÊ", "HA THE"]
    if any(k in name_upper or k in desc_upper for k in lv_indicators):
        return "Hạ thế"
        
    # Fallback to low voltage
    return "Hạ thế"

def clean_project_code(desc):
    """
    Extracts SCL project code from transaction description.
    Tolerates typographical shifts (e.g. mapping VTAD2606001XUẤT -> VTAD2606001).
    """
    if not desc:
        return None
    # Standardize VTDA typo to VTAD
    desc_std = str(desc).upper().replace("VTDA", "VTAD")
    m = re.search(r'(VTAD\d{7})', desc_std)
    if m:
        return m.group(1)
    m_gen = re.search(r'(VTAD\d+)', desc_std)
    if m_gen:
        return m_gen.group(1)
    return None

def generate_voltage_separation_data(import_records, export_records):
    """
    Filters SCL records, separates them by Low/Medium Voltage, 
    standardizes SCL project names, aggregates amounts by Month, Project Code, and Voltage.
    Subtracts SCL returns (Imports SCL returns) from exports.
    Returns a DataFrame.
    """
    rows = []
    
    # Process Exports (XUAT)
    for r in export_records:
        # Space-insensitive SCL check
        desc_clean = r["desc"].replace(" ", "").upper() if r["desc"] else ""
        voucher_clean = r["voucher"].replace(" ", "").upper() if r["voucher"] else ""
        if "SCL" not in desc_clean and "SCL" not in voucher_clean and "VTAD" not in desc_clean and "VTAD" not in voucher_clean and "VTDA" not in desc_clean and "VTDA" not in voucher_clean:
            continue
            
        proj_code = clean_project_code(r["desc"])
        if not proj_code:
            proj_code = clean_project_code(r["voucher"])
            
        # Fallback: if voucher contains .VH4. and it's SCL, it belongs to VTAD2606001
        if not proj_code and ".VH4." in (r["voucher"] or ""):
            proj_code = "VTAD2606001"
            
        if not proj_code:
            continue
            
        vol = classify_voltage(r["code"], r["name"], r["desc"])
        
        # Get Month
        m = None
        if isinstance(r["date"], (datetime.datetime, datetime.date)):
            m = r["date"].month
            
        rows.append({
            "tháng": m,
            "project_code": proj_code,
            "voltage": vol,
            "amount": r["amount"]
        })
        
    # Process Imports (NHAP - SCL Returns & VTTH)
    for r in import_records:
        # Space-insensitive SCL check
        desc_clean = r["desc"].replace(" ", "").upper() if r["desc"] else ""
        voucher_clean = r["voucher"].replace(" ", "").upper() if r["voucher"] else ""
        if "SCL" not in desc_clean and "SCL" not in voucher_clean and "VTAD" not in desc_clean and "VTAD" not in voucher_clean and "VTDA" not in desc_clean and "VTDA" not in voucher_clean:
            continue
            
        proj_code = clean_project_code(r["desc"])
        if not proj_code:
            proj_code = clean_project_code(r["voucher"])
            
        # Fallback: if voucher contains .VH4. and it's SCL, it belongs to VTAD2606001
        if not proj_code and ".VH4." in (r["voucher"] or ""):
            proj_code = "VTAD2606001"
            
        if not proj_code:
            continue
            
        vol = classify_voltage(r["code"], r["name"], r["desc"])
        
        # Get Month
        m = None
        if isinstance(r["date"], (datetime.datetime, datetime.date)):
            m = r["date"].month
            
        itype = classify_import_type(r["desc"], r["code"])
        
        # Only Hoàn Trả (unused exported material returns) reduces net exported material expense.
        # VTTH (recovered tháo gỡ old materials) is NOT subtracted from exported material expense.
        if itype == "HOAN_TRA":
            rows.append({
                "tháng": m,
                "project_code": proj_code,
                "voltage": vol,
                "amount": -r["amount"] # Subtract return amount
            })
        
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["tháng", "project_code", "voltage", "amount"])
        
    # Group by Month, Project Code, and Voltage Level, and sum amounts
    df_grouped = df.groupby(["tháng", "project_code", "voltage"])["amount"].sum().reset_index()
    return df_grouped

def write_to_voltage_template(df_grouped, template_path, output_path, pm_data=None):
    """
    Generates a multi-sheet Excel workbook based on Tách PP-BL.xlsx template.
    Creates a worksheet for each month present, writes project voltage separation lines,
    inserts live Excel formulas, and cell-by-cell copies all font styles, colors, and borders
    from blueprints of rows 12, 13, and 14 in the template.
    
    If pm_data is provided (from parse_pm_092), refund rows (Hoàn nhập) are added
    per project, classified by voltage level.
    """
    wb_template = openpyxl.load_workbook(template_path, data_only=False)
    sheet_blueprint = wb_template.active # Sheet1 acts as blueprint
    blueprint_name = sheet_blueprint.title
    
    # 1. Capture exact cell style blueprints from rows 12, 13, and 14
    blueprints = {}
    for row_num in [12, 13, 14]:
        blueprints[row_num] = {}
        for col_idx in range(1, 9):
            cell = sheet_blueprint.cell(row=row_num, column=col_idx)
            blueprints[row_num][col_idx] = {
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "border": copy(cell.border) if cell.border else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format
            }
            
    # Capture header styles of cells C8, B5
    c8_style = {
        "font": copy(sheet_blueprint.cell(row=8, column=3).font),
        "fill": copy(sheet_blueprint.cell(row=8, column=3).fill),
        "border": copy(sheet_blueprint.cell(row=8, column=3).border),
        "alignment": copy(sheet_blueprint.cell(row=8, column=3).alignment),
        "number_format": sheet_blueprint.cell(row=8, column=3).number_format
    }
    
    # 2. Create sheets for each month
    # Group data by month
    months = sorted(df_grouped["tháng"].dropna().unique())
    if not months:
        # Fallback if no months
        months = [datetime.datetime.now().month]
        
    for month in months:
        month_int = int(month)
        sheet_name = f"Tháng {month_int}"
        
        # Replicate/copy worksheet within the same workbook
        sheet_month = wb_template.copy_worksheet(sheet_blueprint)
        sheet_month.title = sheet_name
        
        # Write Month in B5
        sheet_month.cell(row=5, column=2).value = f"GL_............ ngày     /   {month_int:02d}   /2026"
        
        # Filter data for this month
        df_month = df_grouped[df_grouped["tháng"] == month]
        
        # Get unique projects in this month
        projects_in_month = list(sorted(df_month["project_code"].unique()))
        
        # Also include projects that only have refunds in PM_092 for this month
        refund_by_voltage_month = {}
        if pm_data:
            refund_by_voltage_month = get_pm092_refund_by_voltage(pm_data, month_int)
            for rp in refund_by_voltage_month:
                if rp not in projects_in_month:
                    projects_in_month.append(rp)
            projects_in_month = sorted(projects_in_month)
        
        # Delete original rows 12, 13, 14
        original_max = sheet_month.max_row
        if original_max >= 12:
            sheet_month.delete_rows(12, original_max - 11)
            
        # Clean up existing merged cell ranges that start at or below row 12
        ranges_to_remove = []
        for r in list(sheet_month.merged_cells.ranges):
            if r.min_row >= 12:
                ranges_to_remove.append(r)
        for r in ranges_to_remove:
            sheet_month.merged_cells.remove(r)
            
        stt_counter = 1
        curr_row = 12
        
        for proj_code in projects_in_month:
            proj_name = PROJECT_NAMES.get(proj_code, f"{proj_code} - Dự án Sửa chữa lớn")
            df_proj = df_month[df_month["project_code"] == proj_code]
            proj_refunds = refund_by_voltage_month.get(proj_code, {})
            
            proj_start_row = curr_row
            row_count_in_proj = 0
            
            # Write both Medium and Low voltage rows for each project
            for v_idx, vol in enumerate(["Trung thế", "Hạ thế"]):
                df_vol = df_proj[df_proj["voltage"] == vol]
                amt_val = df_vol["amount"].sum() if not df_vol.empty else 0.0
                
                # Column A: STT (Only on first row of project)
                sheet_month.cell(row=curr_row, column=1).value = stt_counter if row_count_in_proj == 0 else None
                # Column B: Project Name (Only on first row)
                sheet_month.cell(row=curr_row, column=2).value = proj_name if row_count_in_proj == 0 else None
                # Column C: Project Code
                sheet_month.cell(row=curr_row, column=3).value = proj_code
                # Column D: Voltage Level
                sheet_month.cell(row=curr_row, column=4).value = vol
                # Column E: Total Cost
                sheet_month.cell(row=curr_row, column=5).value = amt_val
                
                # Column F: Distribution share formula (=E{row}*80.79%)
                sheet_month.cell(row=curr_row, column=6).value = f"=E{curr_row}*80.79%"
                # Column G: Retail share formula (=E{row}-F{row})
                sheet_month.cell(row=curr_row, column=7).value = f"=E{curr_row}-F{curr_row}"
                
                # Apply styles from Row 12 (Trung thế) or Row 13 (Hạ thế) blueprints
                blueprint_row = 12 if vol == "Trung thế" else 13
                bp = blueprints[blueprint_row]
                for col_idx in range(1, 9):
                    cell = sheet_month.cell(row=curr_row, column=col_idx)
                    style = bp.get(col_idx)
                    if style:
                        if style["font"]: cell.font = style["font"]
                        if style["fill"]: cell.fill = style["fill"]
                        if style["border"]: cell.border = style["border"]
                        if style["alignment"]: cell.alignment = style["alignment"]
                        if style["number_format"]: cell.number_format = style["number_format"]
                        
                curr_row += 1
                row_count_in_proj += 1
            
            # Write refund rows (Hoàn nhập) from PM_092
            if proj_refunds:
                for vol_label in ["Trung thế", "Hạ thế", "Chưa phân loại"]:
                    refund_amt = proj_refunds.get(vol_label, 0.0)
                    if refund_amt > 0:
                        neg_amt = -refund_amt
                        
                        sheet_month.cell(row=curr_row, column=1).value = None
                        sheet_month.cell(row=curr_row, column=2).value = None
                        sheet_month.cell(row=curr_row, column=3).value = proj_code
                        sheet_month.cell(row=curr_row, column=4).value = f"Hoàn nhập - {vol_label}"
                        sheet_month.cell(row=curr_row, column=5).value = neg_amt
                        sheet_month.cell(row=curr_row, column=6).value = f"=E{curr_row}*80.79%"
                        sheet_month.cell(row=curr_row, column=7).value = f"=E{curr_row}-F{curr_row}"
                        
                        # Apply Hạ thế style (row 13) for refund rows
                        bp = blueprints[13]
                        for col_idx in range(1, 9):
                            cell = sheet_month.cell(row=curr_row, column=col_idx)
                            style = bp.get(col_idx)
                            if style:
                                if style["font"]: cell.font = style["font"]
                                if style["fill"]: cell.fill = style["fill"]
                                if style["border"]: cell.border = style["border"]
                                if style["alignment"]: cell.alignment = style["alignment"]
                                if style["number_format"]: cell.number_format = style["number_format"]
                        
                        curr_row += 1
                        row_count_in_proj += 1
                
            # Merge STT (Col A) and Project Name (Col B) for all project rows
            proj_end_row = proj_start_row + row_count_in_proj - 1
            if row_count_in_proj > 1:
                sheet_month.merge_cells(start_row=proj_start_row, start_column=1, end_row=proj_end_row, end_column=1)
                sheet_month.merge_cells(start_row=proj_start_row, start_column=2, end_row=proj_end_row, end_column=2)
            
            # Re-apply styling for Col A and B bottom cells to keep borders and fill consistent
            for merge_r in range(proj_start_row + 1, proj_end_row + 1):
                for col_idx in [1, 2]:
                    cell = sheet_month.cell(row=merge_r, column=col_idx)
                    style = blueprints[13].get(col_idx)
                    if style:
                        if style["font"]: cell.font = style["font"]
                        if style["fill"]: cell.fill = style["fill"]
                        if style["border"]: cell.border = style["border"]
                        if style["alignment"]: cell.alignment = style["alignment"]
                        if style["number_format"]: cell.number_format = style["number_format"]
            
            stt_counter += 1
            
        # Write "Tổng cộng" Row at the end
        tot_row = curr_row
        sheet_month.cell(row=tot_row, column=1).value = "Tổng cộng"
        sheet_month.cell(row=tot_row, column=2).value = None
        sheet_month.cell(row=tot_row, column=3).value = None
        sheet_month.cell(row=tot_row, column=4).value = None
        
        # Merge A{tot_row}:D{tot_row}
        sheet_month.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=4)
        
        # Formulas for Totals
        end_data_row = tot_row - 1
        sheet_month.cell(row=tot_row, column=5).value = f"=SUM(E12:E{end_data_row})"
        sheet_month.cell(row=tot_row, column=6).value = f"=SUM(F12:F{end_data_row})"
        sheet_month.cell(row=tot_row, column=7).value = f"=SUM(G12:G{end_data_row})"
        
        # Apply styles from blueprint row 14 (Tổng cộng)
        bp_tot = blueprints[14]
        for col_idx in range(1, 9):
            cell = sheet_month.cell(row=tot_row, column=col_idx)
            style = bp_tot.get(col_idx)
            if style:
                if style["font"]: cell.font = style["font"]
                if style["fill"]: cell.fill = style["fill"]
                if style["border"]: cell.border = style["border"]
                if style["alignment"]: cell.alignment = style["alignment"]
                if style["number_format"]: cell.number_format = style["number_format"]
                
        # Write sum formula to C8
        c8_cell = sheet_month.cell(row=8, column=3)
        c8_cell.value = f"=E{tot_row}"
        # Apply styling to C8
        if c8_style["font"]: c8_cell.font = c8_style["font"]
        if c8_style["fill"]: c8_cell.fill = c8_style["fill"]
        if c8_style["border"]: c8_cell.border = c8_style["border"]
        if c8_style["alignment"]: c8_cell.alignment = c8_style["alignment"]
        if c8_style["number_format"]: c8_cell.number_format = c8_style["number_format"]
        
    # Delete the blueprint sheet before saving
    wb_template.remove(wb_template[blueprint_name])
    
    try:
        wb_template.save(output_path)
    except PermissionError:
        print(f"Warning: File '{output_path}' is currently open in another program and cannot be saved.", file=sys.stderr)
    except Exception as e:
        print(f"Error saving '{output_path}': {e}", file=sys.stderr)
    return len(months)

def write_detailed_scl_classification(df_scl, template_path, output_path):
    """
    Generates a detailed SCL transaction file with classification column P.
    Loads standard Xuat_Nhap(mau).xlsx template, clears old rows, writes all SCL
    transactions chronologically, applies blueprint styling, and appends the 16th column "PHÂN LOẠI".
    """
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    # 1. Capture the formatting styles from Row 3 to use as a blueprint
    style_blueprint = {}
    for col_idx in range(1, 16):
        cell = sheet.cell(row=3, column=col_idx)
        style_blueprint[col_idx] = {
            "font": copy(cell.font) if cell.font else None,
            "fill": copy(cell.fill) if cell.fill else None,
            "border": copy(cell.border) if cell.border else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
            "number_format": cell.number_format
        }
        
    # Column P style will copy style blueprint from Column O (15)
    style_blueprint[16] = copy(style_blueprint[15])
    # Set column P alignment to center
    if style_blueprint[16]["alignment"]:
        align = copy(style_blueprint[16]["alignment"])
        align.horizontal = "center"
        style_blueprint[16]["alignment"] = align
        
    # Write headers for Column P
    sheet.cell(row=1, column=16).value = "Trung/Hạ thế"
    sheet.cell(row=2, column=16).value = "PHÂN LOẠI"
    
    # Apply header style from O1/O2 to P1/P2
    for r_num in [1, 2]:
        cell_src = sheet.cell(row=r_num, column=15)
        cell_dest = sheet.cell(row=r_num, column=16)
        cell_dest.font = copy(cell_src.font) if cell_src.font else None
        cell_dest.fill = copy(cell_src.fill) if cell_src.fill else None
        cell_dest.border = copy(cell_src.border) if cell_src.border else None
        cell_dest.alignment = copy(cell_src.alignment) if cell_src.alignment else None
        
    # 2. Clear old rows (starting at row 3)
    original_max_row = sheet.max_row
    if original_max_row >= 3:
        sheet.delete_rows(3, original_max_row - 2)
        
    # 3. Write new SCL records with classification
    if not df_scl.empty:
        for r_idx, row_dict in enumerate(df_scl.to_dict('records'), start=3):
            # Classify voltage
            vol = classify_voltage(row_dict["Mã vật tư"], row_dict["Tên vật tư"], row_dict["Diễn giải"])
            
            # Map dictionary keys to standard 15 columns
            col_keys = [
                "Mã vật tư", "Tên vật tư", "ĐVT", "tháng", "năm", "Ngày viết",
                "Số chứng từ (Nhập)", "Số chứng từ (Xuất)", "Diễn giải",
                "Nhập - Số lượng", "Nhập - Đơn giá", "Nhập - Thành tiền",
                "Xuất - Số lượng", "Xuất - Đơn giá", "Xuất - Thành tiền"
            ]
            
            # Write 15 columns
            for c_idx, key in enumerate(col_keys, start=1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                cell_value = row_dict.get(key)
                
                # Format dates nicely
                if isinstance(cell_value, (datetime.datetime, datetime.date)):
                    if isinstance(cell_value, datetime.datetime):
                        cell.value = cell_value.date()
                    else:
                        cell.value = cell_value
                else:
                    cell.value = cell_value
                
                # Apply preserved style blueprint
                style = style_blueprint.get(c_idx)
                if style:
                    if style["font"]: cell.font = style["font"]
                    if style["fill"]: cell.fill = style["fill"]
                    if style["border"]: cell.border = style["border"]
                    if style["alignment"]: cell.alignment = style["alignment"]
                    if style["number_format"]: cell.number_format = style["number_format"]
            
            # Write Column P (PHÂN LOẠI KHÂU TRUNG/HẠ ÁP)
            cell_p = sheet.cell(row=r_idx, column=16)
            cell_p.value = vol
            style_p = style_blueprint.get(16)
            if style_p:
                if style_p["font"]: cell_p.font = style_p["font"]
                if style_p["fill"]: cell_p.fill = style_p["fill"]
                if style_p["border"]: cell_p.border = style_p["border"]
                if style_p["alignment"]: cell_p.alignment = style_p["alignment"]
                if style_p["number_format"]: cell_p.number_format = "@"
                
            # Write Column Q (LOẠI GIAO DỊCH: Xuất / Nhập Hoàn Trả / Nhập VTTH)
            cell_q = sheet.cell(row=r_idx, column=17)
            cell_q.value = row_dict.get("Loại giao dịch", "")
            style_q = style_blueprint.get(17) or style_p
            if style_q:
                if style_q["font"]: cell_q.font = style_q["font"]
                if style_q["fill"]: cell_q.fill = style_q["fill"]
                if style_q["border"]: cell_q.border = style_q["border"]
                if style_q["alignment"]: cell_q.alignment = style_q["alignment"]
                if style_q["number_format"]: cell_q.number_format = "@"
                
    try:
        wb.save(output_path)
    except PermissionError:
        print(f"Warning: File '{output_path}' is currently open in another program and cannot be saved.", file=sys.stderr)
    except Exception as e:
        print(f"Error saving '{output_path}': {e}", file=sys.stderr)
    return len(df_scl)

def classify_pm092_voltage(desc):
    """
    Classifies a PM_092 credit (refund/return) transaction into voltage level
    based on keywords found in the description.
    Returns: "Trung thế", "Hạ thế", or "Chưa phân loại"
    """
    if not desc:
        return "Chưa phân loại"
    
    d = str(desc).upper().replace(" ", "")
    d_original = str(desc).upper()
    
    # --- Hạ thế indicators (check first, more specific) ---
    ha_keywords = ["HẠ THẾ", "HẠ THÊ", "HA THE", "TI HẠ", "TIHẠ"]
    for kw in ha_keywords:
        if kw.replace(" ", "") in d:
            return "Hạ thế"
    
    # --- Trung thế indicators ---
    # "TI TT" = TI trung thế, "TT SCL" = trung thế SCL, "TT thay" = trung thế thay
    trung_keywords = ["TRUNG THẾ", "TRUNG THÊ", "TRUNG THE", "TITT", "TTSCL", "TTTHAY"]
    for kw in trung_keywords:
        if kw.replace(" ", "") in d:
            return "Trung thế"
    
    # "TU" alone (without TI) suggests voltage transformer = Trung thế
    # But "TU, TI" or "TI, TU" is mixed → Chưa phân loại
    # Use regex to find standalone words (handles commas, punctuation)
    import re as _re
    words_in_desc = _re.findall(r'\b[A-ZÀ-Ỹa-zà-ỹ]+\b', d_original)
    has_tu = "TU" in words_in_desc
    has_ti = "TI" in words_in_desc
    
    # Check for comma-separated patterns like "TU, TI" or "TI, TU"
    if has_tu and has_ti:
        return "Chưa phân loại"
    if has_tu and not has_ti:
        return "Trung thế"
    if has_ti and not has_tu:
        # TI alone without voltage indicator → check further
        # If no trung/ha indicator found, classify as Hạ thế (TI hạ thế is more common)
        return "Hạ thế"
    
    return "Chưa phân loại"


def parse_pm_092(file_path):
    """
    Parses PM_092.xlsx (Subledger Account Book for Account 2413).
    Supports both single-month and multi-month files.
    Groups transactions by project code AND month (based on actual transaction dates).
    Returns a dict: {project_code: {"month": latest_month, "debit": total_debit, "credit": total_credit, "net": total_net,
                      "by_month": {month_num: {"debit": sum, "credit": sum, "net": sum}},
                      "credit_lines": [{"month": m, "credit": val, "desc": desc}, ...]}}
    The top-level "month"/"debit"/"credit"/"net" are kept for backward compatibility and represent
    the cumulative totals across all months. The "by_month" sub-dict allows per-month reconciliation.
    The "credit_lines" list stores individual credit transactions with descriptions for voltage classification.
    """
    if isinstance(file_path, str) and not os.path.exists(file_path):
        return {}
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 1. Parse end-month from Row 8 header for backward compatibility
        # Format example: "Từ ngày: 01-05-2026 đến ngày 31-05-2026"
        #            or:  "Từ ngày: 01-01-2026 đến ngày 26-06-2026"
        header_month = None
        row8_val = ws.cell(row=8, column=1).value
        if row8_val:
            match = re.search(r'đến ngày \d{2}-(\d{2})-\d{4}', str(row8_val))
            if match:
                header_month = int(match.group(1))
                
        if not header_month:
            header_month = datetime.datetime.now().month
            
        current_project = None
        # Nested structure with credit_lines for detailed refund tracking
        project_data = {}
        
        for r in range(12, ws.max_row + 1):
            cell_a = ws.cell(row=r, column=1).value
            
            # Detect Project Header
            if cell_a and "Công trình:" in str(cell_a):
                current_project = str(cell_a).split("Công trình:")[1].strip().split("-")[0].strip()
                if current_project not in project_data:
                    project_data[current_project] = {
                        "month": header_month,
                        "debit": 0.0, "credit": 0.0, "net": 0.0,
                        "by_month": {},
                        "credit_lines": []
                    }
                continue
                
            # Detect Detail row (contains date in Column 2)
            date_val = ws.cell(row=r, column=2).value
            if isinstance(date_val, (datetime.datetime, datetime.date)):
                v_code = str(ws.cell(row=r, column=3).value or "").strip()
                desc_val = ws.cell(row=r, column=4).value
                desc_str = str(desc_val).strip() if desc_val else ""
                
                # Filter out internal GL transfer entries ("K/c công trình hoàn thành" / "Kết chuyển")
                if "K/C" in desc_str.upper() or "KẾT CHUYỂN" in desc_str.upper() or v_code.replace("'", "") == "2110":
                    continue

                debit = clean_numeric(ws.cell(row=r, column=5).value or 0.0)
                credit = clean_numeric(ws.cell(row=r, column=6).value or 0.0)
                
                if current_project:
                    # Determine month from the actual transaction date
                    txn_month = date_val.month
                    
                    # Accumulate into per-month buckets
                    if txn_month not in project_data[current_project]["by_month"]:
                        project_data[current_project]["by_month"][txn_month] = {"debit": 0.0, "credit": 0.0, "net": 0.0}
                    
                    project_data[current_project]["by_month"][txn_month]["debit"] += debit
                    project_data[current_project]["by_month"][txn_month]["credit"] += credit
                    
                    # Also accumulate into cumulative totals
                    project_data[current_project]["debit"] += debit
                    project_data[current_project]["credit"] += credit
                    
                    # Store individual credit lines for voltage classification
                    if credit > 0:
                        project_data[current_project]["credit_lines"].append({
                            "month": txn_month,
                            "credit": credit,
                            "desc": desc_str
                        })
                    
        # Calculate Net values
        for proj in project_data:
            project_data[proj]["net"] = project_data[proj]["debit"] - project_data[proj]["credit"]
            for m in project_data[proj]["by_month"]:
                bm = project_data[proj]["by_month"][m]
                bm["net"] = bm["debit"] - bm["credit"]
            
        return project_data
    except Exception as e:
        print(f"Error parsing PM_092: {e}", file=sys.stderr)
        return {}


def get_pm092_refund_by_voltage(pm_data, month):
    """
    Aggregates PM_092 credit (refund/return) amounts by project code and voltage level
    for a given month.
    
    Args:
        pm_data: dict returned by parse_pm_092()
        month: month number to filter
        
    Returns:
        dict: {project_code: {"Trung thế": amount, "Hạ thế": amount, "Chưa phân loại": amount}}
    """
    result = {}
    
    for proj_code, proj_data in pm_data.items():
        credit_lines = proj_data.get("credit_lines", [])
        month_credits = [cl for cl in credit_lines if cl["month"] == month]
        
        if not month_credits:
            continue
            
        voltage_sums = {"Trung thế": 0.0, "Hạ thế": 0.0, "Chưa phân loại": 0.0}
        
        for cl in month_credits:
            vol = classify_pm092_voltage(cl["desc"])
            voltage_sums[vol] += cl["credit"]
            
        result[proj_code] = voltage_sums
        
    return result
